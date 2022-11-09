
from datetime import date
from tkinter.ttk import Style
from functools import reduce

import numpy
import math
import pandas as pd
import plotly.express as px
import streamlit as st
from IPython.display import display
from PIL import Image
# importing the modules
from tabulate import tabulate

st.set_page_config(page_title='Scholar Data')

# st.subheader('Was the tutorial helpful?')
# importing openpyxl module
import openpyxl as xl
from openpyxl import Workbook
flag=0
# streamlit_style = """
# 			<style>
# 			@import url('https://fonts.googleapis.com/css2?family=Roboto:wght@100&display=swap');
#
# 			html, body, [class*="css"]  {
# 			font-family: 'Cursive', sans-serif;
# 			}
# 			</style>
# 			"""
# st.markdown(streamlit_style, unsafe_allow_html=True)







Dept_selection = st.sidebar.radio('Select table options:', ('IFP-Faculties','Journal Publications','Ongoing research scholar','EEE', 'ECE', 'CSE', 'IT', 'RC', 'Consolidated',))
print(Dept_selection)
if Dept_selection == "Journal Publications":
    filename = 'List of Journal Publication - Jan 2022-Dec 2022.xlsx'
    flag=1
elif Dept_selection == "IFP-Faculties":
    filename = 'abc.xlsx'
    flag=2
elif Dept_selection == "Ongoing research scholar":
    filename = 'res.xlsx'
    flag=3

elif Dept_selection == "EEE":
    filename = "EEE PhD guided & guiding by the Supervisors - 16-6-2022.xlsx"

elif Dept_selection == "ECE":
    filename = 'ECE PhD guided & guiding by the Supervisors - 16-6-2022.xlsx'
elif Dept_selection == "IT":
    filename = 'IT PhD guided & guiding by the Supervisors - 16-6-2022.xlsx'
elif Dept_selection == "RC":
    filename = 'RC PhD guided & guiding by the Supervisors - 16-6-2022.xlsx'
elif Dept_selection == "Consolidated":
    filename = 'Consolidated PhD guided & guiding by the Supervisors.xlsx'
elif Dept_selection == "CSE":
    filename = 'CSE PhD guided & guiding by the Supervisors - 16-6-2022.xlsx'


# excel_file='RC PhD guided & guiding by the Supervisors - 16-6-2022.xlsx'
# excel_file1=excel_file[0: excel_file.find(" ")]

# opening the source excel file
# filename ="RC PhD guided & guiding by the Supervisors - 16-6-2022.xlsx"


if 1 == flag:

    df = pd.read_excel(
            io="List of Journal Publication - Jan 2022-Dec 2022.xlsx",
            engine="openpyxl",
            sheet_name="Sheet1",
            usecols="A:Y",
            skiprows=4

    )
    #hirthik
    coauthoro = df['Name of co-author (Outsider)'].head(df.shape[0]).unique().tolist()
    yr1 = df['Year of Publication'].head(df.shape[0]).unique().tolist()
    mon1 = df['Month of Publication'].head(df.shape[0]).unique().tolist()





    thom_count=0
    display(df.columns.values.tolist())
    sno = df['Sl.No'].head(df.shape[0]).unique().tolist()
    firstauthor = df['Name of the First Author'].head(df.shape[0]).tolist()
    paper = df['Paper Title'].head(df.shape[0] ).tolist()
    ppr = df['Paper Title'].head(df.shape[0] ).tolist()
    prp = df['Paper Title'].head(df.shape[0]).tolist()
    coauthor = df['Name of co-author (SSN Faculty)'].head(df.shape[0] ).tolist()
    inter = df['International'].head(df.shape[0]).tolist()
    nat = df['National'].head(df.shape[0]).tolist()
    jn = df['Journal Name'].head(df.shape[0]).tolist()

    mop = df['Impact factor'].head(df.shape[0] - 1).tolist()
    unpaid = df['Unpaid'].head(df.shape[0] - 1).tolist()
    thom = df['Thomson Reuters'].head(df.shape[0]).tolist()
    mon = df['Month of Publication'].head(df.shape[0]).tolist()
    yr = df['Year of Publication'].head(df.shape[0]).tolist()
    ni = df['Non Indexed'].head(df.shape[0] - 1).tolist()
    pid = df['Paper ID (As per SSN Monthly Report)'].head(df.shape[0]).tolist()
    print(firstauthor)




    st.subheader('Number of International and National Publications')
    inter_c=0
    nat_c=0
    for i in inter:
        if(i=='Yes'):
            inter_c+=1
    for i in nat:
        if (i == 'Yes'):
            nat_c+= 1

    st.markdown(f'**International publications count: {(inter_c)}**')
    st.markdown(f'**National publications count: {(nat_c)}**')
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    st.subheader('Number of journals with outsiders as co-authors')
    ni_ca = 0
    for i in coauthoro:
        # if (i == 'Yes'):
        ni_ca += 1

    print(coauthoro)
    print(ni_ca)


    st.markdown(f'*Outsiders as co-authors count: {(ni_ca - 1)}*')

    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    st.subheader('Journals published in a particular year')
    option = st.selectbox(

        'select year',
        (yr1))
    st.write('You selected:', option)

    # name_selection = st.selectbox('Name of the Supervisor:', name)

    aa = 0
    aab = 0
    for i in range(len(yr)):
        if (yr[i] == option):
            aa += 1
        else:
            aab += 1

    if (aa != 0):
        st.markdown(f'*Total journals in {option}: {(aa)}*')
    elif (aab != 0):
        st.markdown(f'*Total journals in {option}: {(aab)}*')

    print(pid)
    st.subheader('Journals published in a particular month')
    mon2 = []
    for i in mon1:
        if i.strip() not in mon2:
            mon2.append(i.strip())
        else:
            pass

    mon4 = []
    for i in mon:
        mon4.append(i.strip())

    option = st.selectbox(

        'Select Month',
        (mon2))
    print(mon2)
    st.write('You selected:', option)
    strOp = str(option)
    print(strOp)
    # name_selection = st.selectbox('Name of the Supervisor:', name)

    aaam = 0
    aabm = 0
    aacm = 0
    aadm = 0
    aaem = 0
    aafm = 0
    aagm = 0
    aahm = 0
    aaim = 0
    aajm = 0
    aakm = 0
    aalm = 0
    for i in range(len(mon4)):
        if (mon4[i] == 'January'):
            aaam += 1
        elif (mon4[i] == 'February'):
            aabm += 1
        elif (mon4[i] == 'March'):
            aacm += 1
        elif (mon4[i] == 'April'):
            aadm += 1
        elif (mon4[i] == 'May'):
            aaem += 1
        elif (mon4[i] == 'June'):
            aafm += 1
        elif (mon4[i] == 'July'):
            aagm += 1
        elif (mon4[i] == 'August'):
            aahm += 1
        elif (mon4[i] == 'September'):
            aaim += 1
        elif (mon4[i] == 'October'):
            aajm += 1
        elif (mon4[i] == 'November'):
            aakm += 1
        elif (mon4[i] == 'December'):
            aalm += 1

    if (option == "January"):
        st.markdown(f'*total journals in {option}: {(aaam)}*')
        # print(aaam)
    elif (option == "February"):
        st.markdown(f'*total journals in {option}: {(aabm)}*')
    elif (option == "March"):
        st.markdown(f'*total journals in {option}: {(aacm)}*')
    elif (option == "April"):
        st.markdown(f'*total journals in {option}: {(aadm)}*')
    elif (option == "May"):
        st.markdown(f'*total journals in {option}: {(aaem)}*')
    elif (option == "June"):
        st.markdown(f'*total journals in {option}: {(aafm)}*')
    elif (option == "July"):
        st.markdown(f'*total journals in {option}: {(aagm)}*')
    elif (option == "August"):
        st.markdown(f'*total journals in {option}: {(aahm)}*')
    elif (option == "September"):
        st.markdown(f'*total journals in {option}: {(aaim)}*')
    elif (option == "October"):
        st.markdown(f'*total journals in {option}: {(aajm)}*')
    elif (option == "November"):
        st.markdown(f'*total journals in {option}: {(aakm)}*')
    elif (option == "December"):
        st.markdown(f'*total journals in {option}: {(aalm)}*')
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    st.subheader('Number of Unpaid journals')

    ni_cup = 0
    for i in unpaid:
        if (i == 'Yes'):
            ni_cup += 1

    print(unpaid)
    print(ni_cup)
    st.markdown(f'*Unpaid Journals count: {(ni_cup + 1)}*')

    st.subheader('Number of non-indexed journals')
    ni_c = 0
    for i in ni:
        if (i =='Yes'):
            ni_c += 1

    print(ni)
    print(ni_c)
    st.markdown(f'**Non Indexed Journals count: {(ni_c)}**')

    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    st.subheader('Paper ID for a particular Paper Title')
    option = st.selectbox(

        'Paper Title',
        (paper))
    st.write('You selected:', option)

    # name_selection = st.selectbox('Name of the Supervisor:', name)

    a = 0
    for i in range(len(paper)):
        if (paper[i] == option):
            a = i

    st.markdown(f'**Paper ID: {(pid[a])}**')
    print(pid)

    st.subheader('Impact factor for a particular Paper Title')


    option = st.selectbox(

        'Paper Title',
        (ppr), key="44")
    st.write('You selected:', option)

    # name_selection = st.selectbox('Name of the Supervisor:', name)

    abc = 0
    for i in range(len(paper)):
        if (ppr[i] == option):
            abc = i

    st.markdown(f'**Impact Factor: {(mop[abc])}**')
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    print(mop)

    st.subheader('Month and Year for a particular Journal')
    option = st.selectbox(

        'Journal Name',
        jn, key='47')
    st.write('You selected:', option)

    # name_selection = st.selectbox('Name of the Supervisor:', name)

    acd = 0
    for j in range(len(paper)):
        if (jn[j] == option):
            acd = j

    st.markdown(f'**Month Of Journal: {(mon[acd])}**')
    st.markdown(f'**Year: {(yr[acd])}**')

    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    print(mon)
    print(yr)

    st.subheader('Paper title and Co Author name for a Particular author')
    option = st.selectbox(

        'Author Name',
        firstauthor, key='48')
    st.write('You selected:', option)

    # name_selection = st.selectbox('Name of the Supervisor:', name)

    dl = 0
    for k in range(len(firstauthor)):
        if (firstauthor[k] == option):
            dl = k


    print(paper)
    print(coauthor)
















    for i in thom:
        if(i=="Yes"):
            thom_count+=1

    st.markdown(f'**Paper Title: {(paper[dl])}**')
    st.markdown(f'**Co-Author: {(coauthor[dl])}**')
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    st.markdown(f'**Count as on date {date.today()}: {len(sno)}**')
    st.subheader('Number of papers under Thompson Reuters Master Journals list')
    st.markdown(f'**Thompson Reuters journals: {(thom_count)}**')

    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    d={}
    for i in mon:
        if i.strip() not in d:
            d[i.strip()] = 1
        else:
            d[i.strip()] += 1
    res = {key: val for key, val in sorted(d.items(), key=lambda ele: ele[1], reverse=True)}
    st.subheader('Month wise publications')
    df_tab=pd.DataFrame(res.items(), columns=['Month', 'Count'])
    st.table(df_tab)


    st.subheader('Impact factor')
    option1 = st.select_slider('Select a number',
                               options=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10'])
    mask = (df['Impact factor'].between(0, int(option1))) & (df['Paper Title'].isin(paper))
    number_of_result = df[mask].shape[0]
    shortlist_name = df[mask].get(['Paper Title'])
    # print(shortlist_name)
    st.markdown(f'*Available Results: {number_of_result}*')

    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    # st.markdown(shortlist_name)
    df[mask].get(['Paper Title', 'Journal Name']).style

    st.subheader('Paper title and authors')
    option = st.selectbox(

        'Paper Title',
        (paper), key='52')
    st.write('You selected:', option)

    # name_selection = st.selectbox('Name of the Supervisor:', name)

    aaa=0
    for i in range(len(paper)):
        if(paper[i] == option):
            aaa=i

    st.markdown(f'**First Author: {(firstauthor[aaa])}**')
    st.markdown(f'**Co author: {(coauthor[aaa])}**')




elif flag==2:

    df = pd.read_excel(
            io="abc.xlsx",
            engine="openpyxl",
            sheet_name="Int_Completed",
            usecols="A:M",
            skiprows=5

    )
    dd = pd.read_excel(
        io="abc.xlsx",
        engine="openpyxl",
        sheet_name="Int_Ongoing",
        usecols="A:K",
        skiprows=4

    )
    amt2 = dd['Amount in Rs. (Lakhs)']
    amt1 = df['Amount in Rs. (Lakhs)']
    dur = df['Duration']
    tit1 = df['Project Title']
    xxx=df['Principal Investigators /                        Co Investigators']
    # Combine the DataFrame-1 & DataFrame-2
    # along horizontal axis using concat() function
    st.subheader('Number of completed internally funded projects')
    st.markdown(f' Projects count: {(len(amt1)-1)}')

    st.subheader('Number of ongoing internally funded projects')

    st.markdown(f'Projects count: {(len(amt2)-1)}')
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)
    st.subheader("COMPLETED PROJECTS")
    df = pd.concat([tit1,xxx,amt1,dur],axis=1)
    df2 = df.tail(-1)
    st.table(df2)
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    st.subheader("ONGOING PROJECTS")

    dur1 = dd['Duration']

    tit2 = dd['Project Title']
    xx = dd['Principal Investigators /                                Co Investigators']
    # Combine the DataFrame-1 & DataFrame-2
    # along horizontal axis using concat() function
    dff = pd.concat([tit2, xx,amt2, dur1], axis=1)
    df3 = dff.tail(-1)
    st.table(df3)


elif flag==3:
    df = pd.read_excel(
        io="res.xlsx",
        engine="openpyxl",
        sheet_name="10-01-2022",
        usecols="A:O",
        skiprows=3

    )
    reg_num = df['Register Number']
    name_of_guide = df['Name of Ph.D. guide'].tolist()
    researcher = df['Name of Ph.D. Researcher']
    status = df['Status of research ']
    x=[]
    y=[]
    z=[]
    prev=''
    for i in range(len(name_of_guide)):
        if(str(name_of_guide[i])!="nan"):
            prev=name_of_guide[i]
            x.append(name_of_guide[i])


        else:
            if(str(researcher[i])!="nan"):
                x.append(prev)
    for i in range(len(researcher)):
        if (str(researcher[i]) == "nan"):
            continue
        else:
            y.append(researcher[i])

    for i in range(len(status)):
        if (str(status[i]) == "nan"):
            continue
        else:
            z.append(status[i])



    d=pd.DataFrame(x)
    u=pd.DataFrame(y)
    o=pd.DataFrame(z)

    dff = pd.concat([d,u,o], axis=1)
    st.subheader("Ongoing Research Scholar Details")
    dff.columns=['Name of guide','PHD researcher','Status of work']
    st.table(dff)
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)

    import openpyxl
    wb = openpyxl.load_workbook(filename)
    sh = wb.active
    st.subheader("Number of full time and part time ")
    c1 = sh['F4']
    time=df[c1.value].tolist()
    full_time=0
    part_time=0
    for i in time:
        if(type(i)!=float):
            if i=='FT' or i[0:2]=='FT':
                full_time+=1
            elif(i=='PT' or 'PT' in i):
                part_time+=1
        else:
            continue
    st.markdown(f'*Number of full time research scholar projects: {full_time}*')
    st.markdown(f'*Number of part time research scholar projects: {part_time}*')
    st.markdown("""<hr style='height:10px;border:none;color:#333;background-color:blue'/>""", unsafe_allow_html=True)
    st.subheader("PHD RESEARCHER, REGISTRATION NUMBER AND DATE OF REGISTRATION")
    reg_no=df["Register Number"]
    aa=[]
    for i in range(len(reg_no)):
        if (str(reg_no[i]) == "nan"):
            continue
        else:
            aa.append(reg_no[i])
    new=[]
    for i in aa:
        if(i=='-'):
            new.append(0)
        else:
            new.append(int(i))


    aaa=pd.DataFrame(new)
    date=df['Date of Reg.']

    bb = []
    for i in range(len(date)):
        if (str(date[i]) == "nan"):
            continue
        else:
            bb.append((str(date[i])[0:11]))
    dates=pd.DataFrame(bb)

    xx=pd.concat([u,aaa,dates], axis=1)
    xx.columns=['Researcher','Registration Number','Date']
    st.table(xx)
    st.markdown("""<hr style='height:10px;border:none;color:white;background-color:blue'/>""", unsafe_allow_html=True)






























else:
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
    filename1 = filename[0: filename.find(" ")]
    print(filename1)
    # opening the destination excel file
    filename2 = filename1 + ".xlsx"
    print(filename2)
    wb = Workbook()
    ws = wb.active
    ws.title = filename1
    wb.save(filename=filename2)

    wb2 = xl.load_workbook(filename2)
    ws2 = wb2.active

    # calculate total number of rows and
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column
    print(mr)
    print(mc)
    # copying the cell values from source
    # excel file to destination excel file
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i, column=j).value = c.value

    # saving the destination excel file
    wb2.save(str(filename2))

    ### --- LOAD DATAFRAME
    excel_file = filename2
    # excel_file='RC PhD guided & guiding by the Supervisors - 16-6-2022.xlsx'
    # excel_file1=excel_file[0: excel_file.find(" ")]
    # excel_file.replace(excel_file, excel_file1)
    # excel_file=excel_file1+'.xlsx'
    # print(excel_file)
    sheet_name = filename1


    df = pd.read_excel(excel_file,
                       sheet_name=sheet_name,
                       usecols="A:G",
                       header=1)

        # --- STREAMLIT SELECTION
    sno = df['Sl No'].head(df.shape[0] - 1).unique().tolist()
    erpcode = df['ERP Code'].head(df.shape[0] - 1).unique().tolist()
    name = df['Name of the Supervisor'].head(df.shape[0] - 1).unique().tolist()
    yor = df['Year of Recognition'].tolist()
    acount = df['No of PhDs Awarded'].head(df.shape[0] - 1).tolist()
    ocount = df['No of PhDs On-going'].head(df.shape[0] - 1).tolist()
    # department = df['Department'].head(df.shape[0] -1).tolist()
    # df.dropna(inplace=True)

    # df.drop(index=df.index[-1],axis=0, inplace=True)
    # print(df.shape[0] -1)
    # df = df.head(df.shape[0] -1)
    st.markdown(f'**Count as on date {date.today()}: {len(sno)}**')
    st.subheader('Number of PhDs Awarded')
    # acount_selection = st.slider('Select a number',
    #                      min_value= 0,
    #                     max_value= int(max(acount)),
    #                     value=((0,int(max(acount)))), step=1)
    option1 = st.select_slider('Select a number',
                               options=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15',
                                        '16', '18', '19', '20'])
    # print(acount_selection)
    st.markdown(f'**Total No. of PhDs Awarded: {sum(acount)}**')
    # print(color)
    mask = (df['No of PhDs Awarded'].between(1, int(option1))) & (df['Name of the Supervisor'].isin(name))
    number_of_result = df[mask].shape[0]
    shortlist_name = df[mask].get(['Name of the Supervisor'])
    # print(shortlist_name)
    st.markdown(f'*Available Results: {number_of_result}*')
    # st.markdown(shortlist_name)
    if (Dept_selection == 'Consolidated'):
        df[mask].get(['Name of the Supervisor', 'No of PhDs Awarded', 'Department']).style
    else:
        df[mask].get(['Name of the Supervisor', 'No of PhDs Awarded']).style
    # display(shortlist_name)

    st.subheader('Number of PhDs On-Going')
    # ocount_selection = st.slider('Select a number:',
    #                       min_value= 0,
    #                     value=(0,int(max(ocount))))
    option2 = st.select_slider('Select a number', options=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'])

    st.markdown(f'**Total No. of PhDs On-Going: {sum(ocount)}**')

    # --- FILTER DATAFRAME BASED ON SELECTION
    mask1 = (df['No of PhDs On-going'].between(1, int(option2))) & (df['Name of the Supervisor'].isin(name))
    number_of_result = df[mask1].shape[0]
    shortlist_name = df[mask1].get(['Name of the Supervisor'])
    # print(shortlist_name)
    st.markdown(f'*Available Results: {number_of_result}*')
    # st.markdown(shortlist_name)
    df[mask1].get(['Name of the Supervisor', 'No of PhDs On-going']).style
    # display(shortlist_name)

    # print(tabulate(df[mask].get(['Name']), headers = 'keys', tablefmt = 'psql'))

    name_selection = st.multiselect('Name of the Supervisor:',
                                    name,
                                    default=name)
    # name_selection = st.selectbox('Name of the Supervisor:', name)

    mask2 = (df['Name of the Supervisor'].isin(name_selection))
    number_of_result = df[mask2].shape[0]
    # shortlist_name1=df[mask2].get(['Name of the Supervisor'])
    # df[mask2].get().style
    # print(shortlist_name)
    # st.markdown(f'*Available Results: {number_of_result}*')
    # st.markdown(shortlist_name)
    table_selection = st.radio('Select table options:',
                               ('ERP Code', 'Year of Recognition', 'No of PhDs Awarded', 'No of PhDs On-going'))
    print(table_selection)
    df[mask2].get(['Name of the Supervisor', table_selection]).style
# st.markdown(f'**{sno} {erpcode}{name}{yor}{acount}{ocount}**')

# df[mask2].get(shortlist_name1)




# --- PLOT PIE CHART
# pie_chart = px.pie(df_participants,
#                 title='Total No. of Participants',
#                 values='Participants',
#                 names='Departments')
#
# st.plotly_chart(pie_chart)
